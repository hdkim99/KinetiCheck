from __future__ import annotations

import csv
import json
import subprocess
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook

from kineticheck.cli import main
from kineticheck.io import read_batch, write_reports

ROOT = Path(__file__).parents[1]


def test_csv_batch_and_xlsx_export(tmp_path: Path) -> None:
    reports = read_batch(ROOT / "examples" / "operating_points.csv")
    assert [report.run_id for report in reports] == ["low-rate", "caution-size", "high-rate"]
    assert [report.status.value for report in reports] == ["PASS", "FAIL", "FAIL"]
    output = tmp_path / "results.xlsx"
    write_reports(reports, output)
    workbook = load_workbook(output, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        assert sheet is not None
        assert sheet.max_row == 4
        headers = [cell.value for cell in next(sheet.iter_rows())]
        assert "c_wp_value" in headers
        assert "c_ah_status" in headers
    finally:
        workbook.close()


def test_xlsx_input_round_trip(tmp_path: Path) -> None:
    with (ROOT / "examples" / "operating_points.csv").open(encoding="utf-8") as handle:
        csv_rows = list(csv.reader(handle))
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    for row in csv_rows:
        sheet.append(row)
    source = tmp_path / "operating-points.xlsx"
    workbook.save(source)
    reports = read_batch(source)
    assert [report.run_id for report in reports] == ["low-rate", "caution-size", "high-rate"]
    output = tmp_path / "result.csv"
    write_reports(reports, output)
    assert "No significant limitation" in output.read_text(encoding="utf-8")


def test_cli_evaluate_validate_batch_and_plot(tmp_path: Path) -> None:
    json_output = tmp_path / "result.json"
    assert (
        main(
            [
                "evaluate",
                str(ROOT / "examples" / "operating_point.json"),
                "--output",
                str(json_output),
            ]
        )
        == 0
    )
    payload = json.loads(json_output.read_text(encoding="utf-8"))
    assert payload["status"] == "PASS"
    assert len(payload["results"]) == 4
    assert main(["validate"]) == 0
    csv_output = tmp_path / "batch.csv"
    assert (
        main(
            ["batch", str(ROOT / "examples" / "operating_points.csv"), "--output", str(csv_output)]
        )
        == 0
    )
    figure = tmp_path / "envelope.svg"
    assert (
        main(
            [
                "plot",
                str(ROOT / "examples" / "operating_points.csv"),
                "--criterion",
                "C_WP",
                "--output",
                str(figure),
            ]
        )
        == 0
    )
    assert "<svg" in figure.read_text(encoding="utf-8")


def test_core_and_cli_do_not_import_gui_or_plot_backends() -> None:
    script = """
import sys
import kineticheck
import kineticheck.cli
assert 'tkinter' not in sys.modules
assert 'matplotlib' not in sys.modules
assert not any(name.startswith(('PyQt', 'PySide')) for name in sys.modules)
"""
    subprocess.run([sys.executable, "-c", script], check=True)
