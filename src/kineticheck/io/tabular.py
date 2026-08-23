"""Batch operating-point import and scientific result export."""

from __future__ import annotations

import csv
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import cast

from openpyxl import Workbook, load_workbook

from kineticheck.application import evaluate_mapping
from kineticheck.models import ScreeningReport


def _present(row: Mapping[str, object], prefix: str) -> bool:
    return any(key.startswith(prefix) and value not in (None, "") for key, value in row.items())


def _float(row: Mapping[str, object], key: str, *, required: bool = True) -> float | None:
    raw = row.get(key)
    if raw in (None, ""):
        if required:
            raise ValueError(f"missing batch column value: {key}")
        return None
    try:
        return float(cast(str | int | float, raw))
    except (TypeError, ValueError) as error:
        raise ValueError(f"batch value {key} must be numeric, got {raw!r}") from error


def _text(row: Mapping[str, object], key: str, *, default: str | None = None) -> str:
    raw = row.get(key)
    if raw in (None, ""):
        if default is None:
            raise ValueError(f"missing batch column value: {key}")
        return default
    return str(raw).strip()


def _q(row: Mapping[str, object], prefix: str, default_unit: str) -> dict[str, object]:
    return {
        "value": _float(row, f"{prefix}_value"),
        "unit": _text(row, f"{prefix}_unit", default=default_unit),
    }


def row_to_mapping(row: Mapping[str, object]) -> dict[str, object]:
    """Translate the documented flat batch schema into an application mapping."""

    rate: dict[str, object] = {
        "value": _float(row, "rate_value"),
        "unit": _text(row, "rate_unit"),
        "basis": _text(row, "rate_basis"),
    }
    if _present(row, "pellet_density_"):
        rate["pellet_density"] = _q(row, "pellet_density", "kg / meter ** 3")
    bed_void = _float(row, "bed_void_fraction", required=False)
    if bed_void is not None:
        rate["bed_void_fraction"] = bed_void

    criteria: dict[str, object] = {}
    if _present(row, "wp_"):
        wp: dict[str, object] = {
            "effective_diffusivity": _q(row, "wp_effective_diffusivity", "meter ** 2 / second"),
            "surface_concentration": _q(row, "wp_surface_concentration", "mol / meter ** 3"),
        }
        threshold = _float(row, "wp_threshold", required=False)
        if threshold is not None:
            wp["threshold"] = threshold
        criteria["weisz_prater"] = wp
    if _present(row, "mm_"):
        mm: dict[str, object] = {
            "reaction_order": _float(row, "mm_reaction_order"),
            "mass_transfer_coefficient": _q(row, "mm_mass_transfer_coefficient", "meter / second"),
            "bulk_concentration": _q(row, "mm_bulk_concentration", "mol / meter ** 3"),
        }
        threshold = _float(row, "mm_threshold", required=False)
        if threshold is not None:
            mm["threshold"] = threshold
        criteria["mears_mass"] = mm
    if _present(row, "mh_"):
        mh: dict[str, object] = {
            "reaction_enthalpy": _q(row, "mh_reaction_enthalpy", "kilojoule / mol"),
            "activation_energy": _q(row, "mh_activation_energy", "kilojoule / mol"),
            "heat_transfer_coefficient": _q(
                row, "mh_heat_transfer_coefficient", "watt / meter ** 2 / kelvin"
            ),
            "bulk_temperature": _q(row, "mh_bulk_temperature", "kelvin"),
        }
        threshold = _float(row, "mh_threshold", required=False)
        if threshold is not None:
            mh["threshold"] = threshold
        criteria["mears_heat"] = mh
    if _present(row, "ah_"):
        ah: dict[str, object] = {
            "reaction_enthalpy": _q(row, "ah_reaction_enthalpy", "kilojoule / mol"),
            "activation_energy": _q(row, "ah_activation_energy", "kilojoule / mol"),
            "effective_thermal_conductivity": _q(
                row, "ah_effective_thermal_conductivity", "watt / meter / kelvin"
            ),
            "surface_temperature": _q(row, "ah_surface_temperature", "kelvin"),
        }
        threshold = _float(row, "ah_threshold", required=False)
        if threshold is not None:
            ah["threshold"] = threshold
        criteria["anderson_heat"] = ah
    return {
        "run_id": _text(row, "run_id"),
        "rate": rate,
        "particle_radius": _q(row, "particle_radius", "meter"),
        "criteria": criteria,
    }


def _csv_rows(path: Path) -> list[dict[str, object]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _xlsx_rows(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError("XLSX workbook contains no active worksheet")
        values = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(values)
        except StopIteration:
            return []
        headers = [str(value).strip() if value is not None else "" for value in raw_headers]
        return [dict(zip(headers, row, strict=False)) for row in values]
    finally:
        workbook.close()


def read_batch(path: Path) -> list[ScreeningReport]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _csv_rows(path)
    elif suffix == ".xlsx":
        rows = _xlsx_rows(path)
    else:
        raise ValueError("batch input must be .csv or .xlsx")
    if not rows:
        raise ValueError("batch input contains no operating points")
    return [evaluate_mapping(row_to_mapping(row)) for row in rows]


def reports_to_rows(reports: Iterable[ScreeningReport]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for report in reports:
        row: dict[str, object] = {
            "run_id": report.run_id,
            "status": report.status.value,
            "interpretation": report.interpretation,
            "volumetric_rate_mol_m3_s": report.volumetric_rate_mol_m3_s,
            "rate_conversion": report.rate_conversion,
            "warnings": " | ".join(report.warnings),
        }
        for result in report.results:
            key = result.symbol.lower()
            row[f"{key}_value"] = result.value
            row[f"{key}_threshold"] = result.threshold
            row[f"{key}_status"] = result.status.value
        rows.append(row)
    return rows


def write_reports(reports: Iterable[ScreeningReport], path: Path) -> None:
    rows = reports_to_rows(reports)
    if not rows:
        raise ValueError("cannot export an empty report collection")
    headers = list(dict.fromkeys(key for row in rows for key in row))
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        return
    if suffix == ".xlsx":
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            raise RuntimeError("new XLSX workbook contains no active worksheet")
        sheet.title = "Screening results"
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header) for header in headers])
        workbook.save(path)
        return
    raise ValueError("batch output must be .csv or .xlsx")
